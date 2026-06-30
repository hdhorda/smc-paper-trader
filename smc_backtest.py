"""
SMC Backtester — FVG + Liquidity Grab Strategy
===============================================
Instrument: NSE cash stocks (1-min OHLCV parquets from Historicalcash/)
Strategy  : Liquidity Grab → FVG formation → FVG retest entry
            SL: beyond FVG boundary  |  TP: 1:2 R:R

Usage:
    # Single symbol, default 1-min, full history
    python smc_backtest.py --symbol RELIANCE

    # Specific timeframe (1 / 3 / 5 / 15 minutes)
    python smc_backtest.py --symbol RELIANCE --tf 5

    # Date-filtered
    python smc_backtest.py --symbol RELIANCE --start 2024-01-01 --end 2024-12-31

    # Session filter (morning / midday / afternoon / full)
    python smc_backtest.py --symbol RELIANCE --session morning

    # Compare all timeframes side-by-side for one symbol
    python smc_backtest.py --symbol RELIANCE --compare-tf

    # Run all FO stocks on a given timeframe
    python smc_backtest.py --all --tf 5

    # Run ALL symbols x ALL timeframes in one shot (<60s target)
    python smc_backtest.py --all-tfs --start 2020-01-01

Output:
    Output/smc_{SYMBOL}_{TF}min.xlsx          (single run)
    Output/smc_{SYMBOL}_tf_comparison.xlsx    (--compare-tf)
    Output/smc_ALL_{TF}min.xlsx               (--all)

Dependencies:
    pip install pandas pyarrow openpyxl
"""

import argparse
import os
from pathlib import Path

import pandas as pd
import numpy as np

# ─────────────────────────────────────────────────────────────
#  CONFIG — tweak these without touching logic
# ─────────────────────────────────────────────────────────────
BASE_DIR        = Path(__file__).parent
DATA_DIR        = BASE_DIR / "Historicalcash"
OUTPUT_DIR      = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Obsidian vault — backtest notes saved here automatically
# Update this path if your vault is in a different location
OBSIDIAN_VAULT  = BASE_DIR.parent.parent / "SMC Price Action Vault"
OBSIDIAN_BT_DIR = OBSIDIAN_VAULT / "Backtests"

# Swing detection: how many bars on each side must be lower/higher
SWING_LOOKBACK  = 10        # N bars left AND right to confirm a swing

# Liquidity Grab: wick must pierce the swing level by at least this %
MIN_WICK_PCT    = 0.05      # 0.05% of price

# FVG: middle candle range must be at least this multiple of EACH neighbor's range
# Document rule: "large relative to the candles on its left and right"
MIN_FVG_SIZE_MULT = 1.2     # middle candle range >= 1.2x the range of both neighbors

# Liquidity Grab: body must be no more than this fraction of the candle's total range
# Document rule: "long wick and a thin candle body, similar to the Dragonfly Doji"
MAX_GRAB_BODY_RATIO = 0.35  # body <= 35% of total candle range

# How many bars AFTER a grab to look for an FVG
FVG_LOOKFORWARD = 20        # bars

# How many bars after FVG forms to watch for a retest entry
ENTRY_LOOKFORWARD = 30      # bars

# Risk:Reward
RR_TARGET       = 2.0

# Session filter  (IST, 24h)
SESSION_TIMES = {
    "morning"  : ("09:15", "11:30"),
    "midday"   : ("11:30", "13:30"),
    "afternoon": ("13:30", "15:15"),
    "full"     : ("09:15", "15:15"),
}
DEFAULT_SESSION = "full"

# Minimum gap to market close — don't enter if < N bars left in session
MIN_BARS_TO_CLOSE = 15

# ── Premium / Discount Zone filter ───────────────────────────
# Bull FVG entry only allowed in Discount zone (price < 50% of swing range)
# Bear FVG entry only allowed in Premium zone  (price > 50% of swing range)
USE_PD_FILTER   = True    # set False to backtest WITHOUT this filter
PD_LOOKBACK     = 100       # bars to look back when finding the bounding swing H/L

# ── CHoCH (Change of Character) filter ───────────────────────
# Bull entries only when last structural break was bullish (CHoCH bull)
# Bear entries only when last structural break was bearish (CHoCH bear)
USE_CHOCH_FILTER = True     # set False to backtest WITHOUT CHoCH filter

# ── Order Block (OB) filter ───────────────────────────────────
# Bull entries only when an active bullish OB overlaps the FVG retest zone.
# Bear entries only when an active bearish OB overlaps the FVG retest zone.
# OB = last opposite-colour candle before a BOS; invalidated when price wicks through.
USE_OB_FILTER = False      # set True to backtest FVG + LiqGrab + OB strategy

# ── BOS (Break of Structure) filter ──────────────────────────
# Differentiates BOS (trend continuation) from CHoCH (trend reversal).
# Bull entries only when the last structural break was a BOS in the bull direction
# (i.e., trend was already bullish before the break — continuation, not reversal).
# Bear entries only when the last structural break was a BOS in the bear direction.
# Note: resets to 'none' after a CHoCH — waits for the next BOS to confirm trend.
USE_BOS_FILTER = False     # set True to backtest FVG + LiqGrab + BOS strategy

# ── Session Filter (entry-time window) ───────────────────────
# Only allow ENTRIES during high-probability NSE session windows (IST):
#   Morning   : 09:15 – 11:30  (institutional order flow, price discovery)
#   Afternoon : 13:30 – 15:15  (institutional book-closing, directional moves)
# Grabs and FVGs can still FORM outside these windows; only the entry bar is gated.
# Lunch hour (11:30–13:30) is excluded — low volume, choppy, stop-hunting range.
USE_SESSION_FILTER = False  # set True to restrict entries to the above windows

# ── Daily HTF Bias filter ─────────────────────────────────────
# Bull entries only when the PREVIOUS day's close > DAILY_BIAS_SMA-day SMA of closes.
# Bear entries only when the PREVIOUS day's close < DAILY_BIAS_SMA-day SMA of closes.
# Aligns intraday setups with higher-timeframe institutional trend direction.
# Based on: ICT 2022 model — "daily chart determines long/short bias for the session."
USE_DAILY_BIAS_FILTER = False  # set True to add daily HTF bias filter
DAILY_BIAS_SMA = 20            # lookback for daily SMA (20 trading days ≈ 1 month)


# ─────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────

VALID_TIMEFRAMES = [1, 3, 5, 15, 60, 120]


def load_1min(symbol: str) -> pd.DataFrame:
    path = DATA_DIR / f"{symbol}_1min.parquet"
    if not path.exists():
        raise FileNotFoundError(f"No parquet found: {path}")
    df = pd.read_parquet(path)
    df["ts"] = pd.to_datetime(df["ts"], utc=False).dt.tz_localize(None)
    df = df.sort_values("ts").reset_index(drop=True)
    df["date"] = df["ts"].dt.date
    return df


def resample_ohlcv(df: pd.DataFrame, tf_minutes: int) -> pd.DataFrame:
    """
    Resample 1-min OHLCV data to any N-minute timeframe.
    Preserves IST session boundaries — bars never cross day boundaries.
    Fully vectorised: single groupby.agg, no Python date loop.
    """
    if tf_minutes == 1:
        return df.copy()

    tmp = df.copy()
    # Floor each timestamp to the tf_minutes bucket (bar open time)
    tmp["_bucket"] = tmp["ts"].dt.floor(f"{tf_minutes}min")

    # Group by (date, bucket) in one vectorised pass — prevents cross-day bars
    out = (
        tmp.groupby(["date", "_bucket"], sort=True)
        .agg(
            tradingsymbol=("tradingsymbol", "first"),
            open         =("open",          "first"),
            high         =("high",          "max"),
            low          =("low",           "min"),
            close        =("close",         "last"),
            volume       =("volume",        "sum"),
        )
        .reset_index()
        .rename(columns={"_bucket": "ts"})
        .drop(columns=["date"])          # will be re-derived from ts
    )
    out["date"] = out["ts"].dt.date
    out = out.sort_values("ts").reset_index(drop=True)
    return out


def apply_session_filter(df: pd.DataFrame, session: str) -> pd.DataFrame:
    start_t, end_t = SESSION_TIMES.get(session, SESSION_TIMES["full"])
    # Integer minute comparison — avoids creating 600K strings via strftime
    sh, sm = int(start_t[:2]), int(start_t[3:])
    eh, em = int(end_t[:2]),   int(end_t[3:])
    start_mins = sh * 60 + sm
    end_mins   = eh * 60 + em
    ts_mins = df["ts"].dt.hour * 60 + df["ts"].dt.minute
    mask = (ts_mins >= start_mins) & (ts_mins <= end_mins)
    return df[mask].reset_index(drop=True)


def detect_swings(df: pd.DataFrame, n: int = SWING_LOOKBACK) -> pd.DataFrame:
    """
    Mark swing highs and swing lows.
    A bar at index i is a swing high if its high is the highest in [i-n, i+n].
    Similarly for swing low.
    Uses a rolling window approach — O(N) via pandas rolling.
    """
    highs = df["high"]
    lows  = df["low"]

    roll_max = highs.rolling(2 * n + 1, center=True, min_periods=2 * n + 1).max()
    roll_min = lows.rolling(2 * n + 1, center=True, min_periods=2 * n + 1).min()

    df = df.copy()
    df["is_swing_high"] = highs == roll_max
    df["is_swing_low"]  = lows  == roll_min
    return df


def detect_liquidity_grabs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Bullish Liquidity Grab:
      - Candle wicks below a recent swing low
      - Close snaps BACK ABOVE the swing low (long bottom wick)
      - wick_size >= MIN_WICK_PCT of price

    Bearish Liquidity Grab:
      - Candle wicks above a recent swing high
      - Close snaps BACK BELOW the swing high (long top wick)
    """
    df = df.copy()
    df["grab_bull"] = False
    df["grab_bear"] = False
    df["grab_bull_level"] = np.nan
    df["grab_bear_level"] = np.nan

    # Build forward-filled swing level arrays for speed
    # "Most recent swing low seen so far" at each bar
    swing_low_vals  = df.loc[df["is_swing_low"],  "low"].reindex(df.index)
    swing_high_vals = df.loc[df["is_swing_high"], "high"].reindex(df.index)

    swing_low_ff  = swing_low_vals.ffill()
    swing_high_ff = swing_high_vals.ffill()

    min_wick   = df["close"] * (MIN_WICK_PCT / 100)
    candle_rng = df["high"] - df["low"]
    body       = (df["close"] - df["open"]).abs()
    # Thin-body filter: body <= MAX_GRAB_BODY_RATIO of total candle range
    # Document: "long wick and a thin candle body, similar to Dragonfly/Gravestone Doji"
    thin_body  = body <= (candle_rng * MAX_GRAB_BODY_RATIO)

    # Bullish grab: low wicks below swing low, close snaps back above, thin body
    bull_mask = (
        (df["low"] < swing_low_ff) &
        (df["close"] > swing_low_ff) &
        ((swing_low_ff - df["low"]) >= min_wick) &
        thin_body
    )

    # Bearish grab: high wicks above swing high, close snaps back below, thin body
    bear_mask = (
        (df["high"] > swing_high_ff) &
        (df["close"] < swing_high_ff) &
        ((df["high"] - swing_high_ff) >= min_wick) &
        thin_body
    )

    df.loc[bull_mask, "grab_bull"] = True
    df.loc[bull_mask, "grab_bull_level"] = swing_low_ff[bull_mask]
    df.loc[bear_mask, "grab_bear"] = True
    df.loc[bear_mask, "grab_bear_level"] = swing_high_ff[bear_mask]

    return df


def detect_fvgs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Bullish FVG  (3-candle pattern):
      - candle[i] is large green candle
      - candle[i-1].high < candle[i+1].low  (gap exists)
      - fvg_bottom = candle[i-1].high, fvg_top = candle[i+1].low
      - body ratio filter: (close-open)/(high-low) >= MIN_FVG_BODY_PCT

    Bearish FVG:
      - candle[i] is large red candle
      - candle[i-1].low > candle[i+1].high
      - fvg_top = candle[i-1].low, fvg_bottom = candle[i+1].high
    """
    df = df.copy()
    df["fvg_bull"]   = False
    df["fvg_bear"]   = False
    df["fvg_top"]    = np.nan
    df["fvg_bottom"] = np.nan

    o, h, l, c = df["open"], df["high"], df["low"], df["close"]
    rng = (h - l).replace(0, np.nan)

    # Document rule: middle candle must be LARGE RELATIVE TO BOTH NEIGHBORS
    # "large relative to the candles on its left and right"
    left_rng  = rng.shift(1)    # range of candle[i-1]
    right_rng = rng.shift(-1)   # range of candle[i+1]
    large_vs_neighbors = (rng >= left_rng * MIN_FVG_SIZE_MULT) & \
                         (rng >= right_rng * MIN_FVG_SIZE_MULT)

    # Bullish FVG at candle i:
    #   - candle[i] is green AND large relative to neighbors
    #   - candle[i-1].high < candle[i+1].low  (gap exists — no overlap)
    #   - fvg_bottom = candle[i-1].high, fvg_top = candle[i+1].low
    bull_gap  = l.shift(-1) > h.shift(1)   # candle[i+1].low > candle[i-1].high
    bull_mask = bull_gap & (c > o) & large_vs_neighbors

    df.loc[bull_mask, "fvg_bull"]   = True
    df.loc[bull_mask, "fvg_bottom"] = h.shift(1)[bull_mask]   # left candle's high
    df.loc[bull_mask, "fvg_top"]    = l.shift(-1)[bull_mask]  # right candle's low

    # Bearish FVG at candle i:
    #   - candle[i] is red AND large relative to neighbors
    #   - candle[i-1].low > candle[i+1].high  (gap exists — no overlap)
    #   - fvg_top = candle[i-1].low, fvg_bottom = candle[i+1].high
    bear_gap  = h.shift(-1) < l.shift(1)   # candle[i+1].high < candle[i-1].low
    bear_mask = bear_gap & (c < o) & large_vs_neighbors

    df.loc[bear_mask, "fvg_bear"]   = True
    df.loc[bear_mask, "fvg_top"]    = l.shift(1)[bear_mask]   # left candle's low
    df.loc[bear_mask, "fvg_bottom"] = h.shift(-1)[bear_mask]  # right candle's high

    return df


# ─────────────────────────────────────────────────────────────
#  CORE BACKTEST ENGINE
# ─────────────────────────────────────────────────────────────

def detect_choch(df: pd.DataFrame) -> pd.DataFrame:
    """
    Change of Character (CHoCH) / Break of Structure (BOS) bias.

    Tracks the most recent structural break direction:
      Bull bias: close breaks above the most recent forward-filled swing high
      Bear bias: close breaks below the most recent forward-filled swing low

    The bias is shifted forward by 1 bar so entries can only use confirmed
    information (no same-bar look-ahead).

    Adds column:
      choch_bias : 'bull' | 'bear' | 'none'
    """
    df = df.copy()

    # Forward-filled most recent confirmed swing levels
    last_sh = df.loc[df["is_swing_high"], "high"].reindex(df.index).ffill()
    last_sl = df.loc[df["is_swing_low"],  "low"].reindex(df.index).ffill()

    # Structural breaks: close crosses above last swing high (bull) or below last swing low (bear)
    break_bull = (df["close"] > last_sh) & last_sh.notna()
    break_bear = (df["close"] < last_sl) & last_sl.notna()

    # Encode: bull=1.0, bear=-1.0, neither=0.0  (bear wins on simultaneous)
    event = pd.Series(0.0, index=df.index)
    event[break_bull & ~break_bear] =  1.0
    event[break_bear]               = -1.0   # bear overrides both

    # Forward-fill the last non-zero event to carry bias across bars
    bias_num = event.replace(0.0, np.nan).ffill().fillna(0.0)

    # Shift by 1: bias from bar i is available to trade on bar i+1
    bias_shifted = bias_num.shift(1).fillna(0.0)

    df["choch_bias"] = bias_shifted.map({1.0: "bull", -1.0: "bear", 0.0: "none"})
    return df


def detect_bos_choch(df: pd.DataFrame) -> pd.DataFrame:
    """
    Differentiates BOS (Break of Structure / trend continuation) from
    CHoCH (Change of Character / trend reversal).

    Tracks a running trend state:
      - First structural break in either direction → CHoCH (establishes trend)
      - Subsequent breaks in the SAME direction   → BOS  (continues trend)
      - Breaks in the OPPOSITE direction           → CHoCH (reverses trend)

    After a CHoCH, bos_bias resets to 'none' until the next BOS fires.
    This means USE_BOS_FILTER only fires on trend-continuation setups,
    not immediately after reversals (those are S1 / CHoCH-filter setups).

    Optimised: only loops through actual break events (~500 per 576K bars),
    not every bar — ~100x faster than a naive per-bar loop.

    Adds column:
      bos_bias : 'bos_bull' | 'bos_bear' | 'none'  (shifted 1 bar — no look-ahead)
    """
    df = df.copy()
    last_sh = df.loc[df["is_swing_high"], "high"].reindex(df.index).ffill().to_numpy(dtype=float)
    last_sl = df.loc[df["is_swing_low"],  "low"].reindex(df.index).ffill().to_numpy(dtype=float)
    close   = df["close"].to_numpy(dtype=float)
    n       = len(close)

    # Vectorised: find bars where a structural break occurred
    broke_up   = (~np.isnan(last_sh)) & (close > last_sh)
    broke_down = (~np.isnan(last_sl)) & (close < last_sl)
    break_idxs = np.where(broke_up | broke_down)[0]

    # Single loop over break events only (~500 breaks vs 576K bars)
    bias      = np.zeros(n, dtype=float)
    trend     = 0       # 0=unknown, 1=bull, -1=bear
    last_bias = 0.0
    prev      = 0

    for i in break_idxs:
        bias[prev:i] = last_bias            # fill gap before this break
        bu = bool(broke_up[i]) and not bool(broke_down[i])
        bd = bool(broke_down[i])
        if bu:
            if trend == 1: last_bias = 1.0  # BOS bull
            else:          last_bias = 0.0; trend = 1   # CHoCH bull → reset
        elif bd:
            if trend == -1: last_bias = -1.0 # BOS bear
            else:           last_bias = 0.0; trend = -1  # CHoCH bear → reset
        bias[i] = last_bias
        prev = i + 1
    bias[prev:] = last_bias                 # fill tail

    bos_shifted = pd.Series(bias, index=df.index).shift(1).fillna(0.0)
    df["bos_bias"] = bos_shifted.map({1.0: "bos_bull", -1.0: "bos_bear", 0.0: "none"})
    return df


def detect_order_blocks(df: pd.DataFrame, lookback: int = 20) -> pd.DataFrame:
    """
    Detect ICT Order Blocks — fully vectorised, ~0.15s per 576K bars.

    Bull OB = last bearish candle before a bullish BOS (close > last swing high).
    Bear OB = last bullish candle before a bearish BOS (close < last swing low).
    OB is invalidated when price wicks through its zone.

    Columns added (shifted 1 bar — no look-ahead):
      ob_bull_top / ob_bull_bot  : active bullish OB zone (nan if none)
      ob_bear_top / ob_bear_bot  : active bearish OB zone (nan if none)
    """
    df = df.copy()
    n   = len(df)
    idx = np.arange(n, dtype=float)

    open_arr  = df["open"].to_numpy(dtype=float)
    close_arr = df["close"].to_numpy(dtype=float)
    high_arr  = df["high"].to_numpy(dtype=float)
    low_arr   = df["low"].to_numpy(dtype=float)

    # ── 1. BOS detection (vectorised) ───────────────────────────
    sh_prices = df.loc[df["is_swing_high"], "high"].reindex(df.index)
    sl_prices = df.loc[df["is_swing_low"],  "low"].reindex(df.index)
    last_sh = sh_prices.ffill().to_numpy(dtype=float)
    last_sl = sl_prices.ffill().to_numpy(dtype=float)
    bull_bos = (close_arr > last_sh) & ~np.isnan(last_sh)
    bear_bos = (close_arr < last_sl) & ~np.isnan(last_sl)

    # ── 2. "Most-recent opposite-colour candle" lookup (vectorised) ─
    bearish_idx = pd.Series(np.where(close_arr < open_arr, idx, np.nan)).ffill().to_numpy()
    bullish_idx = pd.Series(np.where(close_arr > open_arr, idx, np.nan)).ffill().to_numpy()

    ob_bull_top_raw = np.full(n, np.nan)
    ob_bull_bot_raw = np.full(n, np.nan)
    ob_bear_top_raw = np.full(n, np.nan)
    ob_bear_bot_raw = np.full(n, np.nan)

    bull_bos_i = np.where(bull_bos)[0]
    if len(bull_bos_i):
        prev = (bull_bos_i - 1).clip(0)
        j_f  = bearish_idx[prev]                          # float index of last bearish candle
        ok   = (~np.isnan(j_f)) & (j_f >= bull_bos_i - lookback)
        j    = j_f[ok].astype(int)
        ob_bull_top_raw[j] = np.maximum(open_arr[j], close_arr[j])
        ob_bull_bot_raw[j] = np.minimum(open_arr[j], close_arr[j])

    bear_bos_i = np.where(bear_bos)[0]
    if len(bear_bos_i):
        prev = (bear_bos_i - 1).clip(0)
        j_f  = bullish_idx[prev]
        ok   = (~np.isnan(j_f)) & (j_f >= bear_bos_i - lookback)
        j    = j_f[ok].astype(int)
        ob_bear_top_raw[j] = np.maximum(open_arr[j], close_arr[j])
        ob_bear_bot_raw[j] = np.minimum(open_arr[j], close_arr[j])

    # ── 3. Forward-fill with invalidation (single O(N) Cython loop via numba-free trick) ─
    # Encode: build a "reset mask" where price wicks through the current OB bottom/top.
    # We process bull and bear OBs separately using a cumulative group approach.
    def _ffill_with_invalidation(top_raw, bot_raw, invalidate_arr, check_fn):
        """
        check_fn(invalidate_val, bot) -> True when OB should be cleared.
        Uses group-based pandas cummax/cummin trick where possible; falls back to
        a compact Python loop operating only on OB-event bars (much fewer than N).
        """
        active_top = np.full(n, np.nan)
        active_bot = np.full(n, np.nan)
        ob_event_i = np.where(~np.isnan(top_raw))[0].tolist()
        if not ob_event_i:
            return active_top, active_bot

        cur_top = cur_bot = np.nan
        prev    = 0
        for ei in ob_event_i:
            # Scan from prev to ei to check for invalidation
            if not np.isnan(cur_bot):
                inv = np.where(check_fn(invalidate_arr[prev:ei], cur_bot))[0]
                if len(inv):
                    inv_i = prev + inv[0]
                    active_top[prev:inv_i] = cur_top
                    active_bot[prev:inv_i] = cur_bot
                    cur_top = cur_bot = np.nan
                    # Fill from inv_i to ei as nan (already nan)
                    prev = inv_i
                else:
                    active_top[prev:ei] = cur_top
                    active_bot[prev:ei] = cur_bot
            prev = ei
            # New OB at ei overrides
            cur_top = top_raw[ei]; cur_bot = bot_raw[ei]
            active_top[ei] = cur_top; active_bot[ei] = cur_bot
            prev = ei + 1

        # Fill from last OB event to end
        if not np.isnan(cur_bot) and prev < n:
            inv = np.where(check_fn(invalidate_arr[prev:n], cur_bot))[0]
            if len(inv):
                inv_i = prev + inv[0]
                active_top[prev:inv_i] = cur_top
                active_bot[prev:inv_i] = cur_bot
            else:
                active_top[prev:n] = cur_top
                active_bot[prev:n] = cur_bot
        return active_top, active_bot

    bull_check = lambda arr, bot: arr < bot
    bear_check = lambda arr, top: arr > top

    btop, bbot = _ffill_with_invalidation(ob_bull_top_raw, ob_bull_bot_raw, low_arr,  bull_check)
    rtop, rbot = _ffill_with_invalidation(ob_bear_top_raw, ob_bear_bot_raw, high_arr, bear_check)

    # ── 4. Shift by 1 (no look-ahead) ───────────────────────────
    def _shift1(a):
        out = np.roll(a, 1); out[0] = np.nan; return out

    df["ob_bull_top"] = _shift1(btop); df["ob_bull_bot"] = _shift1(bbot)
    df["ob_bear_top"] = _shift1(rtop); df["ob_bear_bot"] = _shift1(rbot)
    return df


def compute_daily_bias_map(df_1min: pd.DataFrame) -> dict:
    """
    Compute daily HTF bias from the full (pre-session-filter) 1-min DataFrame.
    Returns {date: 'bull' | 'bear' | 'none'}.

    Logic: for each trading date D, bias = 'bull' if close[D-1] > SMA20[D-1], else 'bear'.
    Uses strictly PREVIOUS day's data — zero lookahead.
    First DAILY_BIAS_SMA days return 'none' (insufficient history).
    """
    daily_close = df_1min.groupby("date")["close"].last()
    sma20 = daily_close.rolling(DAILY_BIAS_SMA, min_periods=DAILY_BIAS_SMA).mean()
    bias_map: dict = {}
    dates = list(daily_close.index)
    for i, d in enumerate(dates):
        if i == 0:
            bias_map[d] = "none"
            continue
        prev_close = daily_close.iloc[i - 1]
        prev_sma   = sma20.iloc[i - 1]
        if pd.isna(prev_sma):
            bias_map[d] = "none"
        elif prev_close > prev_sma:
            bias_map[d] = "bull"
        else:
            bias_map[d] = "bear"
    return bias_map


def apply_daily_bias(df: pd.DataFrame, bias_map: dict) -> pd.DataFrame:
    """Map precomputed daily_bias values onto each bar by its date."""
    df = df.copy()
    df["daily_bias"] = df["date"].map(bias_map).fillna("none")
    return df


def compute_htf_signal_map(df_htf: "pd.DataFrame", signal_type: str) -> "dict":
    """
    Build a {ts -> active_direction} map from an already-processed HTF DataFrame.
    Reads existing columns (choch_bias / bos_bias / ob_bull_top+ob_bear_top) —
    does NOT re-run detection, so it is fast (pure column read + forward-fill).
    signal_type: 'bos', 'choch', 'ob'
    """
    import pandas as _pd
    import numpy as _np

    df = df_htf  # already has all detection columns from resample_and_prep

    htf_map = {}
    active = None

    if signal_type == 'choch':
        # choch_bias: 'bull' | 'bear' | 'none'
        col = 'choch_bias' if 'choch_bias' in df.columns else None
        for ts, val in zip(df['ts'], df[col] if col else [None]*len(df)):
            if val in ('bull', 'bear'):
                active = val
            htf_map[ts] = active

    elif signal_type == 'bos':
        # bos_bias: 'bos_bull' | 'bos_bear' | 'none'
        col = 'bos_bias' if 'bos_bias' in df.columns else 'choch_bias'
        for ts, val in zip(df['ts'], df[col]):
            if val == 'bos_bull':
                active = 'bull'
            elif val == 'bos_bear':
                active = 'bear'
            elif val in ('bull', 'bear'):   # fallback if col is choch_bias
                active = val
            htf_map[ts] = active

    elif signal_type == 'ob':
        # OB: active bull OB = ob_bull_top not NaN; active bear OB = ob_bear_top not NaN
        bull_col = 'ob_bull_top' if 'ob_bull_top' in df.columns else None
        bear_col = 'ob_bear_top' if 'ob_bear_top' in df.columns else None
        for i, row in enumerate(df.itertuples(index=False)):
            ts = row.ts
            has_bull = bull_col and _pd.notna(getattr(row, bull_col, _np.nan))
            has_bear = bear_col and _pd.notna(getattr(row, bear_col, _np.nan))
            if has_bull and not has_bear:
                active = 'bull'
            elif has_bear and not has_bull:
                active = 'bear'
            # if both or neither: keep previous active
            htf_map[ts] = active

    else:
        raise ValueError(f"Unknown HTF signal_type: {signal_type}")

    return htf_map


def get_htf_direction_at(ts, htf_map: dict, htf_ts_sorted: list) -> "str | None":
    """
    Given a LTF entry timestamp ts, find the most recent HTF candle at or before ts
    and return its active HTF signal direction.
    Uses bisect for O(log n) lookup.
    """
    import bisect as _bisect
    idx = _bisect.bisect_right(htf_ts_sorted, ts) - 1
    if idx < 0:
        return None
    htf_ts = htf_ts_sorted[idx]
    return htf_map.get(htf_ts, None)


def add_pd_zones(df: pd.DataFrame) -> pd.DataFrame:
    """
    Vectorised Premium/Discount zone labelling -- O(N), computed once
    before the backtest loop (replaces per-entry get_pd_zone calls).

    Adds two columns:
      pd_zone    : 'premium' | 'discount' | 'equilibrium' | 'unknown'
      equilibrium: (cum_day_high + cum_day_low) / 2
    """
    df = df.copy()
    # ts.normalize() returns int64-backed Timestamps — 3x faster groupby than Python date objects
    _dk = df["ts"].dt.normalize()
    df["_cum_day_high"] = df.groupby(_dk)["high"].cummax()
    df["_cum_day_low"]  = df.groupby(_dk)["low"].cummin()
    df["equilibrium"]   = (df["_cum_day_high"] + df["_cum_day_low"]) / 2
    day_bar_num = df.groupby(_dk).cumcount()      # 0-based
    valid = day_bar_num >= 9                        # need >=10 bars
    df["pd_zone"] = "unknown"
    df.loc[valid & (df["close"] > df["equilibrium"]),  "pd_zone"] = "premium"
    df.loc[valid & (df["close"] < df["equilibrium"]),  "pd_zone"] = "discount"
    df.loc[valid & (df["close"] == df["equilibrium"]), "pd_zone"] = "equilibrium"
    df.drop(columns=["_cum_day_high", "_cum_day_low"], inplace=True)
    return df

def run_backtest(df, symbol, htf_map=None, htf_ts_sorted=None):
    trades      = []
    in_trade    = False
    trade_end_i = -1
    n = len(df)
    df = df.copy()
    df["_bar_num_in_day"] = df.groupby("date").cumcount()
    day_counts = df.groupby("date")["ts"].count()
    df["_bars_in_day"] = df["date"].map(day_counts)
    df["_bars_left"]   = df["_bars_in_day"] - df["_bar_num_in_day"]
    close_arr     = df["close"].to_numpy(dtype=float)
    low_arr       = df["low"].to_numpy(dtype=float)
    high_arr      = df["high"].to_numpy(dtype=float)
    bars_left_arr = df["_bars_left"].to_numpy(dtype=int)
    ts_arr        = df["ts"].to_numpy()
    fvg_top_arr   = df["fvg_top"].to_numpy(dtype=float)
    fvg_bot_arr   = df["fvg_bottom"].to_numpy(dtype=float)
    pd_zone_arr   = df["pd_zone"].to_numpy() if "pd_zone" in df.columns else None
    equil_arr     = df["equilibrium"].to_numpy(dtype=float) if "equilibrium" in df.columns else None
    choch_arr     = df["choch_bias"].to_numpy() if "choch_bias" in df.columns else None
    bos_arr           = df["bos_bias"].to_numpy()    if "bos_bias"    in df.columns else None
    daily_bias_arr    = df["daily_bias"].to_numpy()  if "daily_bias"  in df.columns else None
    # Pre-compute bar times in minutes-since-midnight for fast session checking
    ts_mins_arr = (df["ts"].dt.hour * 60 + df["ts"].dt.minute).to_numpy(dtype=int)
    ob_bull_top_arr = df["ob_bull_top"].to_numpy(dtype=float) if "ob_bull_top" in df.columns else None
    ob_bull_bot_arr = df["ob_bull_bot"].to_numpy(dtype=float) if "ob_bull_bot" in df.columns else None
    ob_bear_top_arr = df["ob_bear_top"].to_numpy(dtype=float) if "ob_bear_top" in df.columns else None
    ob_bear_bot_arr = df["ob_bear_bot"].to_numpy(dtype=float) if "ob_bear_bot" in df.columns else None
    bull_fvg_arr  = np.where(df["fvg_bull"].to_numpy())[0]
    bear_fvg_arr  = np.where(df["fvg_bear"].to_numpy())[0]
    grab_bull_arr  = df["grab_bull"].to_numpy()
    grab_bear_arr  = df["grab_bear"].to_numpy()
    grab_bull_lvl  = df["grab_bull_level"].to_numpy(dtype=float)
    grab_bear_lvl  = df["grab_bear_level"].to_numpy(dtype=float)
    grab_idxs = np.where(grab_bull_arr | grab_bear_arr)[0]
    for i in grab_idxs:
        if in_trade and i <= trade_end_i:
            continue
        in_trade = False
        for direction in ("bull", "bear"):
            if direction == "bull":
                if not grab_bull_arr[i]: continue
                grab_level = grab_bull_lvl[i]
                fvg_arr    = bull_fvg_arr
            else:
                if not grab_bear_arr[i]: continue
                grab_level = grab_bear_lvl[i]
                fvg_arr    = bear_fvg_arr
            grab_ts = ts_arr[i]
            j = int(np.searchsorted(fvg_arr, i + 1))
            if j >= len(fvg_arr) or fvg_arr[j] > i + FVG_LOOKFORWARD:
                continue
            fvg_found  = int(fvg_arr[j])
            fvg_top    = fvg_top_arr[fvg_found]
            fvg_bottom = fvg_bot_arr[fvg_found]
            entry_end   = min(fvg_found + ENTRY_LOOKFORWARD, n - 1)
            entry_found = -1
            for k in range(fvg_found + 2, entry_end + 1):
                if bars_left_arr[k] < MIN_BARS_TO_CLOSE:
                    break
                prev_close = close_arr[k - 1]
                kc = close_arr[k]; kl = low_arr[k]; kh = high_arr[k]
                if direction == "bull":
                    if kc < fvg_bottom: break
                    if prev_close > fvg_top and kl <= fvg_top and kc >= fvg_bottom:
                        entry_found = k; break
                else:
                    if kc > fvg_top: break
                    if prev_close < fvg_bottom and kh >= fvg_bottom and kc <= fvg_top:
                        entry_found = k; break
            if entry_found < 0:
                continue
            # ── Session filter: entry bar must be within IST morning or afternoon window ──
            if USE_SESSION_FILTER:
                em = int(ts_mins_arr[entry_found])
                morning_ok   = (9 * 60 + 15 <= em <= 11 * 60 + 30)  # 09:15–11:30
                afternoon_ok = (13 * 60 + 30 <= em <= 15 * 60 + 15)  # 13:30–15:15
                if not (morning_ok or afternoon_ok):
                    continue
            # ── Daily HTF bias filter: align with previous-day's trend direction ────────
            if USE_DAILY_BIAS_FILTER and daily_bias_arr is not None:
                db = daily_bias_arr[entry_found]
                if direction == "bull" and db != "bull":
                    continue
                if direction == "bear" and db != "bear":
                    continue
            # ── HTF signal pre-filter (S5 MTF) ──────────────────────────
            if htf_map is not None and htf_ts_sorted is not None:
                entry_time = df["ts"].iloc[entry_found]
                htf_dir = get_htf_direction_at(entry_time, htf_map, htf_ts_sorted)
                if htf_dir is None or htf_dir != direction:
                    continue
            pd_zone     = pd_zone_arr[entry_found] if pd_zone_arr is not None else "unknown"
            equilibrium = float(equil_arr[entry_found]) if equil_arr is not None else None
            if USE_PD_FILTER:
                if direction == "bull" and pd_zone not in ("discount", "unknown"):
                    continue
                if direction == "bear" and pd_zone not in ("premium",  "unknown"):
                    continue
            if USE_CHOCH_FILTER and choch_arr is not None:
                choch_bias = choch_arr[entry_found]
                if direction == "bull" and choch_bias not in ("bull", "none"):
                    continue
                if direction == "bear" and choch_bias not in ("bear", "none"):
                    continue
            if USE_BOS_FILTER and bos_arr is not None:
                bos_bias = bos_arr[entry_found]
                if direction == "bull" and bos_bias != "bos_bull":
                    continue
                if direction == "bear" and bos_bias != "bos_bear":
                    continue
            if USE_OB_FILTER and ob_bull_top_arr is not None:
                # Require FVG zone to overlap with active Order Block
                if direction == "bull":
                    ob_top = ob_bull_top_arr[entry_found]
                    ob_bot = ob_bull_bot_arr[entry_found]
                    if np.isnan(ob_top) or fvg_top < ob_bot or fvg_bottom > ob_top:
                        continue   # no OB or no overlap
                else:
                    ob_top = ob_bear_top_arr[entry_found]
                    ob_bot = ob_bear_bot_arr[entry_found]
                    if np.isnan(ob_top) or fvg_top < ob_bot or fvg_bottom > ob_top:
                        continue
            entry_price = close_arr[entry_found]
            if direction == "bull":
                sl_price = fvg_bottom; risk = entry_price - sl_price
                if risk <= 0: continue
                tp_price = entry_price + risk * RR_TARGET
            else:
                sl_price = fvg_top; risk = sl_price - entry_price
                if risk <= 0: continue
                tp_price = entry_price - risk * RR_TARGET
            # Scan only to end of current day — bars_left caps the window
            eod_bars   = int(bars_left_arr[entry_found])
            search_end = min(entry_found + 1 + eod_bars, n)
            if search_end <= entry_found + 1:
                continue
            fut_lo = low_arr[entry_found + 1 : search_end]
            fut_hi = high_arr[entry_found + 1 : search_end]
            fut_bl = bars_left_arr[entry_found + 1 : search_end]
            if direction == "bull":
                sl_hit_s = fut_lo <= sl_price
                tp_hit_s = fut_hi >= tp_price
            else:
                sl_hit_s = fut_hi >= sl_price
                tp_hit_s = fut_lo <= tp_price
            eod_hit_s = fut_bl <= 1
            hits_s = sl_hit_s | tp_hit_s | eod_hit_s
            if not hits_s.any():
                continue
            exit_rel = int(hits_s.argmax())
            exit_i   = entry_found + 1 + exit_rel
            is_sl = bool(sl_hit_s[exit_rel]); is_tp = bool(tp_hit_s[exit_rel])
            if is_sl and not is_tp:
                exit_reason = "SL";  exit_price = sl_price
            elif is_tp and not is_sl:
                exit_reason = "TP";  exit_price = tp_price
            elif is_sl and is_tp:
                exit_reason = "SL";  exit_price = sl_price
            else:
                exit_reason = "EOD"; exit_price = float(close_arr[exit_i])
            exit_ts   = ts_arr[exit_i]
            pnl_pts   = (exit_price - entry_price) if direction == "bull" else (entry_price - exit_price)
            pnl_pct   = pnl_pts / entry_price * 100
            rr_actual = pnl_pts / risk if risk > 0 else 0
            win       = exit_reason == "TP" or pnl_pts > 0
            trades.append({
                "symbol": symbol, "direction": direction,
                "grab_ts": grab_ts, "grab_level": round(grab_level, 2),
                "fvg_ts": ts_arr[fvg_found], "fvg_top": round(fvg_top, 2), "fvg_bottom": round(fvg_bottom, 2),
                "entry_ts": ts_arr[entry_found], "entry_price": round(entry_price, 2),
                "sl_price": round(sl_price, 2), "tp_price": round(tp_price, 2), "risk_pts": round(risk, 2),
                "exit_ts": exit_ts, "exit_price": round(exit_price, 2), "exit_reason": exit_reason,
                "pnl_pts": round(pnl_pts, 2), "pnl_pct": round(pnl_pct, 4), "rr_actual": round(rr_actual, 2),
                "win": win, "pd_zone": pd_zone, "equilibrium": equilibrium,
                "choch_bias":    choch_arr[entry_found]      if choch_arr      is not None else "none",
                "daily_bias":    daily_bias_arr[entry_found] if daily_bias_arr is not None else "none",
            })
            in_trade    = True
            trade_end_i = exit_i
            break
    return pd.DataFrame(trades)

def compute_summary(trades: pd.DataFrame, symbol: str) -> pd.DataFrame:
    if trades.empty:
        return pd.DataFrame([{"symbol": symbol, "note": "No trades found"}])

    total   = len(trades)
    wins    = trades["win"].sum()
    losses  = total - wins
    wr      = wins / total * 100

    gross_profit = trades.loc[trades["pnl_pts"] > 0, "pnl_pts"].sum()
    gross_loss   = trades.loc[trades["pnl_pts"] < 0, "pnl_pts"].abs().sum()
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else np.inf

    avg_win  = trades.loc[trades["win"],  "pnl_pts"].mean()
    avg_loss = trades.loc[~trades["win"], "pnl_pts"].mean()
    expectancy = (wr / 100 * avg_win) + ((1 - wr / 100) * avg_loss)

    bull_trades = trades[trades["direction"] == "bull"]
    bear_trades = trades[trades["direction"] == "bear"]

    tp_exits  = (trades["exit_reason"] == "TP").sum()
    sl_exits  = (trades["exit_reason"] == "SL").sum()
    eod_exits = (trades["exit_reason"] == "EOD").sum()

    # Premium / Discount breakdown (only present when pd_zone column exists)
    disc_trades = prem_trades = disc_wr = prem_wr = "N/A"
    if "pd_zone" in trades.columns:
        disc_df = trades[trades["pd_zone"] == "discount"]
        prem_df = trades[trades["pd_zone"] == "premium"]
        disc_trades = len(disc_df)
        prem_trades = len(prem_df)
        disc_wr = round(disc_df["win"].mean() * 100, 1) if disc_trades else 0
        prem_wr = round(prem_df["win"].mean() * 100, 1) if prem_trades else 0

    summary = {
        "symbol"          : symbol,
        "total_trades"    : total,
        "wins"            : int(wins),
        "losses"          : int(losses),
        "win_rate_pct"    : round(wr, 1),
        "profit_factor"   : round(profit_factor, 2),
        "avg_win_pts"     : round(avg_win, 2),
        "avg_loss_pts"    : round(avg_loss, 2),
        "expectancy_pts"  : round(expectancy, 2),
        "total_pnl_pts"   : round(trades["pnl_pts"].sum(), 2),
        "total_pnl_pct"   : round(trades["pnl_pct"].sum(), 2),
        "avg_rr_actual"   : round(trades["rr_actual"].mean(), 2),
        "bull_trades"     : len(bull_trades),
        "bull_win_rate"   : round(bull_trades["win"].mean() * 100, 1) if len(bull_trades) else 0,
        "bear_trades"     : len(bear_trades),
        "bear_win_rate"   : round(bear_trades["win"].mean() * 100, 1) if len(bear_trades) else 0,
        "tp_exits"        : int(tp_exits),
        "sl_exits"        : int(sl_exits),
        "eod_exits"       : int(eod_exits),
        "pd_filter_on"    : USE_PD_FILTER,
        "discount_trades" : disc_trades,
        "discount_wr_pct" : disc_wr,
        "premium_trades"  : prem_trades,
        "premium_wr_pct"  : prem_wr,
        "date_from"       : str(trades["entry_ts"].min().date()),
        "date_to"         : str(trades["entry_ts"].max().date()),
    }
    return pd.DataFrame([summary])


# ─────────────────────────────────────────────────────────────
#  OBSIDIAN NOTE GENERATOR
# ─────────────────────────────────────────────────────────────

def generate_obsidian_note(trades: pd.DataFrame, summary: pd.DataFrame,
                           symbol: str, tf_minutes: int,
                           start: str, end: str, session: str) -> None:
    """
    Write a uniform backtest result note to the Obsidian vault.
    File: Backtests/BT_{SYMBOL}_{TF}min_{DATE}.md
    """
    try:
        OBSIDIAN_BT_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        print("  [Obsidian] Could not create vault folder — skipping note.")
        return

    from datetime import date as _date
    today     = str(_date.today())
    note_name = f"BT_{symbol}_{tf_minutes}min_{today}.md"
    note_path = OBSIDIAN_BT_DIR / note_name

    s = summary.iloc[0] if not summary.empty else {}

    # ── Trade distribution by month ───────────────────────────
    monthly_lines = ""
    if not trades.empty:
        trades["month"] = pd.to_datetime(trades["entry_ts"]).dt.to_period("M")
        monthly = trades.groupby("month").agg(
            trades_n   = ("win", "count"),
            wins_n     = ("win", "sum"),
            pnl        = ("pnl_pts", "sum"),
        )
        monthly["wr"] = (monthly["wins_n"] / monthly["trades_n"] * 100).round(1)
        rows = []
        for period, row in monthly.iterrows():
            rows.append(
                f"| {period} | {int(row.trades_n)} | {int(row.wins_n)} "
                f"| {row.wr}% | {row.pnl:.2f} |"
            )
        monthly_lines = "\n".join(rows)

    # ── Sample winning and losing trades ──────────────────────
    def trade_table(df_sub, n=5):
        if df_sub.empty:
            return "_None_"
        cols = ["entry_ts", "direction", "entry_price", "sl_price",
                "tp_price", "exit_reason", "pnl_pts", "rr_actual"]
        rows = ["| " + " | ".join(cols) + " |",
                "| " + " | ".join(["---"] * len(cols)) + " |"]
        for _, r in df_sub.head(n).iterrows():
            rows.append("| " + " | ".join(str(r[c]) for c in cols) + " |")
        return "\n".join(rows)

    wins_table   = trade_table(trades[trades["win"]]  if not trades.empty else trades)
    losses_table = trade_table(trades[~trades["win"]] if not trades.empty else trades)

    # ── Assemble note ─────────────────────────────────────────
    pf    = s.get("profit_factor", "N/A")
    wr    = s.get("win_rate_pct",  "N/A")
    exp   = s.get("expectancy_pts","N/A")
    pnl   = s.get("total_pnl_pts", "N/A")
    total = s.get("total_trades",  0)
    pf_emoji = "✅" if isinstance(pf, (int, float)) and pf >= 1.0 else "❌"

    note = f"""# Backtest: FVG + Liquidity Grab — {symbol} — {tf_minutes}min

**Date run:** {today}
**Tags:** #backtest #smc #fvg #liquidity-grab #{symbol.lower()} #{tf_minutes}min

---

## Strategy Reference
- [[FVG - Fair Value Gap]]
- [[Liquidity Grabs]]
- [[Strategy Index]]

---

## Test Parameters

| Parameter | Value |
|---|---|
| Symbol | {symbol} (NSE Cash) |
| Timeframe | {tf_minutes}-minute |
| Date Range | {start or "Full history"} → {end or "Latest"} |
| Session | {session} (09:15–15:15 IST) |
| R:R Target | 1:{RR_TARGET} |
| Swing Lookback | {SWING_LOOKBACK} bars each side |

---

## Entry & Exit Checklist

### Bias Signal — Liquidity Grab
- [x] Candle wicks beyond most recent swing high (bearish) or swing low (bullish)
- [x] Close snaps back above SSL / below BSL on same candle
- [x] Wick size ≥ {MIN_WICK_PCT}% of price
- [x] Body ≤ {int(MAX_GRAB_BODY_RATIO*100)}% of candle range (Dragonfly/Gravestone Doji shape)

### Entry Zone — FVG Formation (within {FVG_LOOKFORWARD} bars of grab)
- [x] 3-candle imbalance: candle[i-1].high < candle[i+1].low (bull) or candle[i-1].low > candle[i+1].high (bear)
- [x] Middle candle is green (bull) or red (bear)
- [x] Middle candle range ≥ {MIN_FVG_SIZE_MULT}x the range of BOTH neighbors
- [x] FVG must form AFTER the liquidity grab (not before)

### Confluence Filter — Premium / Discount Zone ({"ON" if USE_PD_FILTER else "OFF"})
- [x] Range basis: **current trading day's high and low** (ICT intraday standard)
- [x] Equilibrium = (Day High + Day Low) / 2
- [x] Discount zone: entry price < Equilibrium  →  valid for **bull** FVG setups (buying cheap)
- [x] Premium zone:  entry price > Equilibrium  →  valid for **bear** FVG setups (selling expensive)

### Entry Trigger — FVG Retest (within {ENTRY_LOOKFORWARD} bars of FVG)
- [x] Price approaches the FVG from the correct side (above for bull / below for bear)
- [x] Candle low enters the FVG zone: low ≤ fvg_top (bull) / high ≥ fvg_bottom (bear)
- [x] Candle closes WITHIN or BEYOND the zone: close ≥ fvg_bottom (bull) / close ≤ fvg_top (bear)
- [x] Entry price = close of retest bar
- [x] At least {MIN_BARS_TO_CLOSE} bars remaining in session

### Stop Loss
- [x] SL = fvg_bottom (bull) / fvg_top (bear)
- [x] Risk = entry_price − SL (bull) / SL − entry_price (bear)

### Take Profit
- [x] TP = entry_price + risk × {RR_TARGET} (bull) / entry_price − risk × {RR_TARGET} (bear)

### Invalidation (FVG cancelled before entry)
- [x] Bull FVG invalid: close < fvg_bottom → stop scanning
- [x] Bear FVG invalid: close > fvg_top → stop scanning

### Exit
- TP hit → WIN
- SL hit → LOSS
- EOD (session close) → exit at close, outcome by PnL sign

---

## Results Summary

| Metric | Value |
|---|---|
| Total Trades | {total} |
| Win Rate | {wr}% |
| Profit Factor | {pf} {pf_emoji} |
| Expectancy | {exp} pts/trade |
| Total PnL | {pnl} pts |
| Avg Win | {s.get("avg_win_pts", "N/A")} pts |
| Avg Loss | {s.get("avg_loss_pts", "N/A")} pts |
| Avg R:R Achieved | {s.get("avg_rr_actual", "N/A")} |
| Bull Trades / WR | {s.get("bull_trades", "N/A")} / {s.get("bull_win_rate", "N/A")}% |
| Bear Trades / WR | {s.get("bear_trades", "N/A")} / {s.get("bear_win_rate", "N/A")}% |
| P/D Filter | {"ON" if USE_PD_FILTER else "OFF"} (lookback {PD_LOOKBACK} bars) |
| Discount Zone Trades / WR | {s.get("discount_trades", "N/A")} / {s.get("discount_wr_pct", "N/A")}% |
| Premium Zone Trades / WR | {s.get("premium_trades", "N/A")} / {s.get("premium_wr_pct", "N/A")}% |
| Bear Trades / WR | {s.get("bear_trades", "N/A")} / {s.get("bear_win_rate", "N/A")}% |
| TP Exits | {s.get("tp_exits", "N/A")} |
| SL Exits | {s.get("sl_exits", "N/A")} |
| EOD Exits | {s.get("eod_exits", "N/A")} |
| Period | {s.get("date_from", "N/A")} → {s.get("date_to", "N/A")} |

---

## Monthly Breakdown

| Month | Trades | Wins | Win Rate | PnL (pts) |
|---|---|---|---|---|
{monthly_lines if monthly_lines else "| — | — | — | — | — |"}

---

## Sample Winning Trades

{wins_table}

---

## Sample Losing Trades

{losses_table}

---

## Observations & Notes

> _Add manual observations here after reviewing the trade log._

- [ ] Bull or bear setups performing better?
- [ ] Any time-of-day clustering in wins/losses?
- [ ] FVG size or location patterns?
- [ ] P/D Zone filter performing as expected on bull vs bear setups?
- [ ] Any time-of-day clustering in wins/losses?
- [ ] Next confluence layer to test: [[CHoCH - Change of Character]] bias filter

---

## Next Steps
- [ ] Add [[CHoCH - Change of Character]] as higher-TF trend bias filter
- [ ] Test with `USE_PD_FILTER = False` to measure isolated P/D zone contribution
- [ ] Test on additional symbols: HDFCBANK, TCS, INFY
- [ ] Compare results with [[Strategy Index]]
"""

    note_path.write_text(note, encoding="utf-8")
    print(f"  [Obsidian] Note saved → {note_path}")


# ─────────────────────────────────────────────────────────────
#  SAVE TO EXCEL
# ─────────────────────────────────────────────────────────────

def save_results(trades: pd.DataFrame, summary: pd.DataFrame, symbol: str, tf_minutes: int = 1):
    out_path = OUTPUT_DIR / f"smc_{symbol}_{tf_minutes}min.xlsx"

    # ── Step 1: write data and close the file cleanly ─────────
    # Formatting happens in Step 2 so an openpyxl error can never
    # corrupt the file (missing EOCD / truncated ZIP).
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        if not trades.empty:
            trades.to_excel(writer, sheet_name="Trades", index=False)
    # File is now fully written and closed.

    # ── Step 2: reopen for formatting (non-fatal if it fails) ──
    if not trades.empty:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.load_workbook(out_path)
            ws = wb["Trades"]
            green = PatternFill("solid", fgColor="C6EFCE")
            red   = PatternFill("solid", fgColor="FFC7CE")
            win_col = list(trades.columns).index("win")  # 0-based
            n_cols  = len(trades.columns)
            for row_idx in range(2, len(trades) + 2):
                win_val = ws.cell(row=row_idx, column=win_col + 1).value
                fill = green if win_val else red
                for col in range(1, n_cols + 1):
                    ws.cell(row=row_idx, column=col).fill = fill
            # Bold header row
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Auto-width columns (cap at 30)
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)
            wb.save(out_path)
        except Exception as fmt_err:
            print(f"  [warn] Formatting skipped: {fmt_err}")

    print(f"  Saved -> {out_path}")
    return out_path


# ─────────────────────────────────────────────────────────────
#  SINGLE SYMBOL RUN
# ─────────────────────────────────────────────────────────────

def run_symbol(symbol, start=None, end=None, session=DEFAULT_SESSION, tf_minutes=1):
    print(f"\n{'='*60}")
    print(f"  Symbol : {symbol}  |  TF: {tf_minutes}min  |  Session: {session}")
    df = load_1min(symbol)
    print(f"  Loaded : {len(df):,} 1-min bars  ({df['ts'].min().date()} -> {df['ts'].max().date()})")
    if start:
        df = df[df["ts"] >= pd.Timestamp(start)]
    if end:
        df = df[df["ts"] <= pd.Timestamp(end)]
    if tf_minutes > 1:
        df = resample_ohlcv(df, tf_minutes)
        print(f"  Resampled to {tf_minutes}min: {len(df):,} bars")
    df = apply_session_filter(df, session)
    print(f"  After session filter: {len(df):,} bars")
    df = df.reset_index(drop=True)
    df = detect_swings(df)
    df = detect_liquidity_grabs(df)
    df = detect_fvgs(df)
    df = detect_choch(df)          # CHoCH bias — structural break direction
    df = add_pd_zones(df)          # vectorised P/D labelling — O(N), done once
    print(f"  Swing H/L : {df['is_swing_high'].sum()} / {df['is_swing_low'].sum()}")
    print(f"  Grabs B/S : {df['grab_bull'].sum()} bull  /  {df['grab_bear'].sum()} bear")
    print(f"  FVGs  B/S : {df['fvg_bull'].sum()} bull  /  {df['fvg_bear'].sum()} bear")
    import time as _time
    _t0 = _time.perf_counter()
    trades = run_backtest(df, symbol)
    print(f"  Backtest   : {_time.perf_counter()-_t0:.2f}s")
    summary = compute_summary(trades, symbol)
    summary.insert(1, "tf_minutes", tf_minutes)
    print(f"\n  Trades     : {len(trades)}")
    if not trades.empty:
        print(f"  Win rate   : {summary['win_rate_pct'].values[0]}%")
        print(f"  Profit F   : {summary['profit_factor'].values[0]}")
        print(f"  Expectancy : {summary['expectancy_pts'].values[0]} pts")
        print(f"  Total PnL  : {summary['total_pnl_pts'].values[0]} pts")
    save_results(trades, summary, symbol, tf_minutes)
    generate_obsidian_note(trades, summary, symbol, tf_minutes, start, end, session)
    return trades, summary


# ─────────────────────────────────────────────────────────────
#  TIMEFRAME COMPARISON FOR ONE SYMBOL
# ─────────────────────────────────────────────────────────────

def run_tf_comparison(symbol, start=None, end=None, session=DEFAULT_SESSION):
    print(f"\n{'='*60}")
    print(f"  TF comparison -- {symbol}  ({start or 'all'} to {end or 'all'})")
    all_summaries = []
    all_trades    = []
    for tf in VALID_TIMEFRAMES:
        trades, summary = run_symbol(symbol, start=start, end=end, session=session, tf_minutes=tf)
        all_summaries.append(summary)
        if not trades.empty:
            trades["tf_minutes"] = tf
            all_trades.append(trades)
    comparison = pd.concat(all_summaries, ignore_index=True)
    out_path = OUTPUT_DIR / f"smc_{symbol}_tf_comparison.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        comparison.to_excel(writer, sheet_name="TF Comparison", index=False)
        if all_trades:
            pd.concat(all_trades, ignore_index=True).to_excel(writer, sheet_name="All Trades", index=False)
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.sheets["TF Comparison"]
        hf = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        green = PatternFill("solid", fgColor="C6EFCE")
        red   = PatternFill("solid", fgColor="FFC7CE")
        pf_col = list(comparison.columns).index("profit_factor") + 1
        for row_idx in range(2, len(comparison) + 2):
            try:
                pf_val = float(ws.cell(row=row_idx, column=pf_col).value or 0)
                fill = green if pf_val >= 1.0 else red
                for col in range(1, len(comparison.columns) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill
            except (TypeError, ValueError):
                pass
    print(f"\n  TF comparison saved -> {out_path}")
    # Write Obsidian comparison note
    try:
        OBSIDIAN_BT_DIR.mkdir(parents=True, exist_ok=True)
        from datetime import date as _date
        today = str(_date.today())
        cmp_path = OBSIDIAN_BT_DIR / f"BT_{symbol}_TF_Comparison_{today}.md"
        rows = []
        for _, r in comparison.iterrows():
            tf   = int(r.get("tf_minutes", 0))
            pf   = r.get("profit_factor", "N/A")
            wr   = r.get("win_rate_pct",  "N/A")
            exp  = r.get("expectancy_pts","N/A")
            tot  = r.get("total_trades",  "N/A")
            pnl  = r.get("total_pnl_pts", "N/A")
            rows.append(f"| {tf}min | {tot} | {wr}% | {pf} | {exp} | {pnl} |")
        pd_status = "ON" if USE_PD_FILTER else "OFF"
        cmp_note = f"""# TF Comparison: FVG + Liquidity Grab -- {symbol}

**Date run:** {today}
**P/D Zone Filter:** {pd_status}
**Tags:** #backtest #smc #tf-comparison #{symbol.lower()}

---

## Summary

| Timeframe | Trades | Win Rate | Profit Factor | Expectancy (pts) | Total PnL (pts) |
|---|---|---|---|---|---|
{chr(10).join(rows)}

> Profit Factor >= 1.5 = edge present. 1.0-1.5 = needs confluence. <1.0 = no standalone edge.

---

## Individual Run Notes

{"".join(f"- [[BT_{symbol}_{int(r.get('tf_minutes',0))}min_{today}]]" + chr(10) for _, r in comparison.iterrows())}
---

## Next Steps
- [ ] Add [[CHoCH - Change of Character]] bias filter to best-performing timeframe
- [ ] Test with USE_PD_FILTER = False to measure isolated P/D zone contribution
- [ ] Run --all across FO universe on best timeframe
"""
        cmp_path.write_text(cmp_note, encoding="utf-8")
        print(f"  [Obsidian] Comparison note saved -> {cmp_path}")
    except Exception as e:
        print(f"  [Obsidian] Could not write comparison note: {e}")
    return comparison


# ─────────────────────────────────────────────────────────────
#  MULTI-SYMBOL RUN
# ─────────────────────────────────────────────────────────────

# Module-level worker — must be top-level so Windows spawn can pickle it
def _run_symbol_worker(args):
    sym, start, end, session, tf_minutes = args
    try:
        trades, summary = run_symbol(sym, start=start, end=end,
                                     session=session, tf_minutes=tf_minutes)
        return trades, summary
    except Exception as e:
        print(f"  ERROR on {sym}: {e}")
        return pd.DataFrame(), pd.DataFrame()


def _run_all_tfs_worker(args):
    import sys, time as _time
    sym, start, end, session, tfs = args
    t0 = _time.perf_counter()
    results = []
    try:
        df_base = load_1min(sym)
        if start: df_base = df_base[df_base["ts"] >= pd.Timestamp(start)]
        if end:   df_base = df_base[df_base["ts"] <= pd.Timestamp(end)]
        # Compute daily bias from FULL data (pre-session-filter) to use true EOD closes
        daily_bias_map = compute_daily_bias_map(df_base) if USE_DAILY_BIAS_FILTER else {}
        # Apply session filter ONCE on 1-min before resampling (avoids 4× strftime)
        df_base = apply_session_filter(df_base, session).reset_index(drop=True)
        for tf in tfs:
            try:
                df = resample_ohlcv(df_base, tf) if tf > 1 else df_base.copy()
                df = detect_swings(df)
                df = detect_liquidity_grabs(df)
                df = detect_fvgs(df)
                df = detect_choch(df)
                df = detect_bos_choch(df)
                df = detect_order_blocks(df)
                df = add_pd_zones(df)
                if USE_DAILY_BIAS_FILTER:
                    df = apply_daily_bias(df, daily_bias_map)
                trades  = run_backtest(df, sym)
                summary = compute_summary(trades, sym)
                summary.insert(1, "tf_minutes", tf)
                if not trades.empty:
                    trades["tf_minutes"] = tf
                results.append((trades, summary))
            except Exception as e:
                print(f"  ERROR {sym} {tf}min: {e}", flush=True)
        elapsed = _time.perf_counter() - t0
        print(f"  OK  {sym:<15} {elapsed:.1f}s", flush=True)
    except Exception as e:
        print(f"  FAIL {sym}: {e}",         flush=True)
    return results


def _init_filters(pd_filter, choch_filter, ob_filter, bos_filter, session_filter, daily_bias_filter):
    """Pool initializer — propagates CLI flag overrides to each worker process (needed on Windows spawn)."""
    global USE_PD_FILTER, USE_CHOCH_FILTER, USE_OB_FILTER, USE_BOS_FILTER
    global USE_SESSION_FILTER, USE_DAILY_BIAS_FILTER
    USE_PD_FILTER          = pd_filter
    USE_CHOCH_FILTER       = choch_filter
    USE_OB_FILTER          = ob_filter
    USE_BOS_FILTER         = bos_filter
    USE_SESSION_FILTER     = session_filter
    USE_DAILY_BIAS_FILTER  = daily_bias_filter


def run_all(start=None, end=None, session=DEFAULT_SESSION, tf_minutes=1):
    import multiprocessing as _mp, time as _time
    symbols = sorted(f.stem.replace("_1min", "") for f in DATA_DIR.glob("*_1min.parquet"))
    n_cores = max(1, _mp.cpu_count() - 1)
    print(f"Found {len(symbols)} symbols  |  TF: {tf_minutes}min  |  Session: {session}")
    print(f"Running on {n_cores} parallel workers...")
    t0 = _time.perf_counter()
    args_list = [(sym, start, end, session, tf_minutes) for sym in symbols]
    with _mp.Pool(n_cores, maxtasksperchild=10) as pool:
        results = pool.map(_run_symbol_worker, args_list)
    elapsed = _time.perf_counter() - t0
    print(f"Completed {len(symbols)} symbols in {elapsed:.1f}s ({elapsed/len(symbols):.2f}s/symbol)")


WATCHLIST_20 = [
    "ABCAPITAL","APLAPOLLO","PFC","RELIANCE",
    "BOSCHLTD","JSWSTEEL","IEX","OIL",
    "LUPIN","GODREJCP","PAGEIND","LODHA",
    "ALKEM","DIXON","JUBLFOOD","ETERNAL",
    "GMRAIRPORT","DABUR","BAJAJFINSV","FORCEMOT",
]

def run_all_tfs(start=None, end=None, session=DEFAULT_SESSION, tfs=None, symbols=None):
    import multiprocessing as _mp, time as _time, sys
    if tfs is None:
        tfs = VALID_TIMEFRAMES
    if symbols is None:
        symbols = sorted(f.stem.replace("_1min", "") for f in DATA_DIR.glob("*_1min.parquet"))
    else:
        symbols = sorted(symbols)
    n_cores = max(1, _mp.cpu_count() - 1)
    total_combos = len(symbols) * len(tfs)
    print(f"Found {len(symbols)} symbols x {len(tfs)} TFs = {total_combos} combinations", flush=True)
    print(f"Running on {n_cores} parallel workers (parquet loaded once per symbol)...", flush=True)
    t0 = _time.perf_counter()
    args_list = [(sym, start, end, session, tfs) for sym in symbols]
    all_results = []
    done = 0
    with _mp.Pool(n_cores, maxtasksperchild=10,
                  initializer=_init_filters,
                  initargs=(USE_PD_FILTER, USE_CHOCH_FILTER, USE_OB_FILTER, USE_BOS_FILTER,
                            USE_SESSION_FILTER, USE_DAILY_BIAS_FILTER)) as pool:
        for sym_result in pool.imap_unordered(_run_all_tfs_worker, args_list):
            all_results.append(sym_result)
            done += 1
            elapsed = _time.perf_counter() - t0
            eta = (elapsed / done) * (len(symbols) - done) if done else 0
            print(f"  Progress: {done}/{len(symbols)} symbols  |  {elapsed:.0f}s elapsed  |  ETA ~{eta:.0f}s", flush=True)
    elapsed = _time.perf_counter() - t0
    print(f"\nCompleted in {elapsed:.1f}s ({elapsed/len(symbols):.2f}s/symbol, {elapsed/total_combos:.2f}s/combo)", flush=True)
    all_trades  = []
    all_summary = []
    for sym_results in all_results:
        for trades, summary in sym_results:
            if not trades.empty:  all_trades.append(trades)
            if not summary.empty: all_summary.append(summary)
    if all_summary:
        combined_summary = pd.concat(all_summary, ignore_index=True)
        combined_trades  = pd.concat(all_trades,  ignore_index=True) if all_trades else pd.DataFrame()
        filters = []
        if USE_PD_FILTER:          filters.append("PD")
        if USE_CHOCH_FILTER:       filters.append("CHOCH")
        if USE_OB_FILTER:          filters.append("OB")
        if USE_BOS_FILTER:         filters.append("BOS")
        if USE_SESSION_FILTER:     filters.append("SESSION")
        if USE_DAILY_BIAS_FILTER:  filters.append("DBIAS")
        suffix = "_" + "_".join(filters) if filters else "_nofilter"
        out_path = OUTPUT_DIR / f"smc_ALL_AllTF{suffix}.xlsx"
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            combined_summary.to_excel(writer, sheet_name="Summary",    index=False)
            if not combined_trades.empty:
                combined_trades.to_excel(writer, sheet_name="All Trades", index=False)
        print(f"Saved -> {out_path}", flush=True)
    else:
        print("WARNING: No results collected. Check for errors above.", flush=True)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SMC FVG + Liquidity Grab Backtester")
    parser.add_argument("--symbol",     type=str)
    parser.add_argument("--all",        action="store_true")
    parser.add_argument("--all-tfs",    action="store_true",
                        help="All symbols x all TFs in one shot (fastest)")
    parser.add_argument("--watchlist",  action="store_true",
                        help="Run only the 20-stock watchlist (fast, ~30s)")
    parser.add_argument("--compare-tf", action="store_true")
    parser.add_argument("--tf",         type=int, default=1, choices=VALID_TIMEFRAMES)
    parser.add_argument("--start",      type=str, default=None)
    parser.add_argument("--end",        type=str, default=None)
    parser.add_argument("--session",    type=str, default=DEFAULT_SESSION,
                        choices=list(SESSION_TIMES.keys()))
    # ── Filter flags (override config-file defaults) ──────────────────
    parser.add_argument("--no-pd-filter",      action="store_true",
                        help="Disable PD Zone filter (overrides USE_PD_FILTER=True in config)")
    parser.add_argument("--no-choch-filter",   action="store_true",
                        help="Disable CHoCH filter (overrides USE_CHOCH_FILTER=True in config)")
    parser.add_argument("--ob-filter",         action="store_true",
                        help="Enable Order Block filter (overrides USE_OB_FILTER=False in config)")
    parser.add_argument("--bos-filter",        action="store_true",
                        help="Enable BOS filter -- trend-continuation only (overrides USE_BOS_FILTER=False)")
    parser.add_argument("--session-filter",    action="store_true",
                        help="Restrict ENTRIES to morning (09:15-11:30 IST) and afternoon (13:30-15:15 IST) windows only")
    parser.add_argument("--daily-bias-filter", action="store_true",
                        help="Only take entries aligned with daily HTF trend (prev-day close vs 20-day SMA)")
    parser.add_argument("--output-dir",        type=str, default=None,
                        help="Override output directory (default: Output/ next to script)")
    args = parser.parse_args()

    # Apply CLI overrides to module-level globals (direct assignment at module scope)
    if args.no_pd_filter:
        USE_PD_FILTER = False
    if args.no_choch_filter:
        USE_CHOCH_FILTER = False
    if args.ob_filter:
        USE_OB_FILTER = True
    if args.bos_filter:
        USE_BOS_FILTER   = True
        USE_CHOCH_FILTER = False   # BOS and CHoCH are mutually exclusive filters
    if args.session_filter:
        USE_SESSION_FILTER = True
    if args.daily_bias_filter:
        USE_DAILY_BIAS_FILTER = True
    if args.output_dir:
        OUTPUT_DIR = Path(args.output_dir)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.all_tfs:
        syms = WATCHLIST_20 if args.watchlist else None
        run_all_tfs(start=args.start, end=args.end, session=args.session, symbols=syms)
    elif args.compare_tf:
        if not args.symbol:
            print("ERROR: --compare-tf requires --symbol")
        else:
            run_tf_comparison(args.symbol.upper(), start=args.start, end=args.end, session=args.session)
    elif args.all:
        run_all(start=args.start, end=args.end, session=args.session, tf_minutes=args.tf)
    elif args.symbol:
        run_symbol(args.symbol.upper(), start=args.start, end=args.end,
                   session=args.session, tf_minutes=args.tf)
    else:
        parser.print_help()
